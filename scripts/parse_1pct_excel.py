#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
parse_1pct_excel.py — Parser laporan KSEI "Kepemilikan Investor di atas 1%"
dari file XLSX (ekspor langsung dari sistem KSEI ke Excel, BUKAN Excel yang
di-print-to-PDF). Alternatif untuk parse_1pct.py, dipakai saat sumbernya
Excel, bukan PDF — mis. saat PDF yang tersedia adalah hasil convert Excel
yang strukturnya sudah tidak bisa diandalkan (spasi kolom/grid tidak
konsisten dari bulan ke bulan).

Kontrak keluaran IDENTIK dengan parse_1pct.py: CSV 12 kolom yang sama,
persis urutan sama, kolom date diformat "DD-Mon-YYYY" sama seperti versi
PDF (mis. "31-Jul-2026") — supaya merge_fakta.py bisa langsung konsumsi
CSV keluaran skrip ini TANPA modifikasi apa pun.

Cara pakai:
    python parse_1pct_excel.py "1 Juli 2026.xlsx" -o ksei_1pct_20260731.csv \
        [--sheet "Sheet 1"] [--report laporan.md] [--baseline-csv ksei_saham_20260703.csv]

Perbedaan pendekatan vs versi PDF:
- Tidak ada masalah "kata menempel" atau "grid tidak ketemu" — struktur
  tabel Excel sudah eksplisit per sel, tidak perlu direkonstruksi dari
  posisi visual.
- Baris header dicari secara DINAMIS (baris pertama yang sel A & B-nya
  persis "DATE" dan "SHARE_CODE"), bukan nomor baris hardcoded — supaya
  tahan kalau paragraf penafian di bulan berikutnya berubah panjang.
- Angka & tanggal dibaca langsung sebagai tipe data Python asli (int/
  float/datetime) dari openpyxl, tidak perlu regex parsing seperti versi
  PDF (yang harus urai "1.234.567" / "12,34" dari teks visual).
- Validasi bawaan PERSIS SAMA dengan parse_1pct.py: TOTAL == SCRIPLESS +
  SCRIP, persentase di (0,100], kode 4-karakter, nama emiten seragam per
  kode, dsb — supaya baku mutu data yang masuk ke fakta.csv konsisten
  antara sumber PDF dan sumber Excel.
"""

import argparse
import csv
import datetime as dt
import json
import sys
from collections import Counter, defaultdict

import openpyxl

# Kolom keluaran & pola validasi disamakan PERSIS dengan parse_1pct.py
# supaya kontrak data konsisten lintas sumber (PDF vs Excel).
import re

CODE_RE = re.compile(r"^[A-Z0-9]{4}$")

MONTH_ABBR = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
}

COLS = [
    "date", "share_code", "issuer_name", "investor_name",
    "investor_classification", "local_foreign", "nationality", "domicile",
    "holdings_scripless", "holdings_scrip", "total_holding_shares", "percentage",
]

EXPECTED_HEADER = [
    "DATE", "SHARE_CODE", "ISSUER_NAME", "INVESTOR_NAME",
    "INVESTOR_CLASSIFICATION", "LOCAL_FOREIGN", "NATIONALITY", "DOMICILE",
    "HOLDINGS_SCRIPLESS", "HOLDINGS_SCRIP", "TOTAL_HOLDING_SHARES", "PERCENTAGE",
]


def date_to_str(d):
    """datetime(2026,7,31) -> '31-Jul-2026' (format sama dgn keluaran parse_1pct.py)."""
    return f"{d.day:02d}-{MONTH_ABBR[d.month]}-{d.year}"


def find_header_row(ws, max_scan=50):
    """Cari baris pertama yang sel A & B-nya persis 'DATE'/'SHARE_CODE'."""
    for i, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan, max_col=2, values_only=True), start=1
    ):
        if row[0] == "DATE" and row[1] == "SHARE_CODE":
            return i
    raise RuntimeError(
        f"Baris header 'DATE'/'SHARE_CODE' tidak ditemukan dalam {max_scan} "
        f"baris pertama. Cek manual posisi header di file ini."
    )


def parse_excel(path, sheet_name=None):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    header_row = find_header_row(ws)
    actual_header = next(
        ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
    )
    # Hanya 12 kolom pertama yang wajib cocok persis. Kolom ekstra di luar
    # itu (mis. kolom ke-13 kosong tanpa header, artefak Excel) ditoleransi
    # ASAL memang kosong di header-nya -- kalau ternyata berlabel, itu bisa
    # jadi kolom baru yang perlu ditinjau, jadi dicatat sbg peringatan
    # (bukan crash) supaya tidak diam-diam kebuang.
    if list(actual_header[:12]) != EXPECTED_HEADER:
        raise RuntimeError(
            f"Header ditemukan di baris {header_row} tapi 12 kolom pertama "
            f"beda dari yang diharapkan.\n  ditemukan: {list(actual_header[:12])}\n"
            f"  diharap  : {EXPECTED_HEADER}"
        )
    extra_header_labels = [v for v in actual_header[12:] if v is not None]

    records = []
    issues = defaultdict(list)
    stats = {
        "sheet": ws.title, "header_row": header_row, "baris_kosong": 0,
        "extra_header_labels": extra_header_labels,
    }

    for rownum, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, max_col=12, values_only=True),
        start=header_row + 1,
    ):
        (date, code, issuer, investor, classif, lf, nat, dom,
         scripless, scrip, total, pct) = row

        if date is None and code is None:
            stats["baris_kosong"] += 1
            continue

        loc = f"baris {rownum}"

        if not isinstance(date, dt.datetime):
            issues["tanggal_tidak_valid"].append(f"{loc}: {date!r}")
            continue
        date_s = date_to_str(date)

        code = (code or "").strip()
        if not CODE_RE.match(code):
            issues["kode_tidak_valid"].append(f"{loc}: {code!r}")
            continue

        if scripless is None or scrip is None or total is None:
            issues["angka_kosong"].append(f"{loc}: {code} scripless={scripless} scrip={scrip} total={total}")
            continue
        scripless, scrip, total = int(scripless), int(scrip), int(total)

        if pct is None:
            issues["pct_kosong"].append(f"{loc}: {code}")
            continue
        pct = float(pct)

        lf = (lf or "").strip()
        if lf not in ("L", "F", ""):
            issues["lf_tidak_valid"].append(f"{loc}: {lf!r}")

        issuer = (issuer or "").strip()
        investor = (investor or "").strip()

        if scripless + scrip != total:
            issues["total_tak_konsisten"].append(
                f"{loc}: {code} {investor!r}: {scripless}+{scrip}!={total}"
            )
        if not (0 < pct <= 100):
            issues["pct_di_luar_batas"].append(f"{loc}: {code} {pct}")
        if not issuer:
            issues["emiten_kosong"].append(f"{loc}: {code}")
        if not investor:
            issues["investor_kosong"].append(f"{loc}: {code}")

        records.append({
            "date": date_s,
            "share_code": code,
            "issuer_name": issuer,
            "investor_name": investor,
            "investor_classification": (classif or "").strip(),
            "local_foreign": lf,
            "nationality": (nat or "").strip(),
            "domicile": (dom or "").strip(),
            "holdings_scripless": scripless,
            "holdings_scrip": scrip,
            "total_holding_shares": total,
            "percentage": pct,
            "_row": rownum,
        })

    by_code = {}
    for r in records:
        by_code.setdefault(r["share_code"], set()).add(r["issuer_name"])
    for k, v in sorted(by_code.items()):
        if len(v) > 1:
            issues["emiten_tidak_seragam"].append(f"{k}: {sorted(v)}")

    return records, issues, stats


def build_report(records, issues, stats, src_path, baseline_rows=None):
    lines = []
    add = lines.append
    dates = Counter(r["date"] for r in records)
    codes = sorted({r["share_code"] for r in records})
    classif = Counter(r["investor_classification"] for r in records)
    lf = Counter(r["local_foreign"] or "(kosong)" for r in records)
    pair_counts = Counter((r["share_code"], r["investor_name"]) for r in records)
    dup_pairs = {p: c for p, c in pair_counts.items() if c > 1}
    scrip_pos = sum(1 for r in records if r["holdings_scrip"] > 0)

    add("# Laporan Validasi Tahap 1 — Parser Excel KSEI 1%")
    add("")
    add(f"Sumber: `{src_path}` (sheet: {stats['sheet']!r}, header di baris {stats['header_row']})")
    add("")
    add("## Hasil ekstraksi")
    add(f"- Baris data: **{len(records)}**")
    add(f"- Emiten unik: **{len(codes)}**")
    add(f"- Investor unik: **{len({r['investor_name'] for r in records})}**")
    add(f"- Tanggal periode: **{dict(dates)}**")
    add(f"- Baris ber-scrip (>0): {scrip_pos}")
    add(f"- Baris kosong dilewati: {stats['baris_kosong']}")
    if stats.get("extra_header_labels"):
        add(f"- **PERHATIAN**: ada label kolom di luar 12 kolom standar (kolom ke-13+): "
            f"{stats['extra_header_labels']} — kolom ini TIDAK dibaca/diikutkan, tinjau manual "
            f"kalau ternyata berisi data penting.")
    add("")
    add("## Distribusi nilai")
    add(f"- Lokal/Asing: {dict(lf.most_common())}")
    add(f"- Klasifikasi investor ({len(classif)} jenis): {dict(classif.most_common())}")
    add(f"- Pasangan (kode, investor) muncul >1×: {len(dup_pairs)}"
        + (f" -> {list(dup_pairs.items())[:10]}" if dup_pairs else ""))
    add("")
    add("## Pemeriksaan integritas")
    ok = True
    for key, label in [
        ("tanggal_tidak_valid", "Tanggal tidak valid"),
        ("kode_tidak_valid", "Kode emiten tidak valid"),
        ("angka_kosong", "Kolom angka kosong"),
        ("pct_kosong", "Persentase kosong"),
        ("lf_tidak_valid", "Nilai Lokal/Asing tak dikenal"),
        ("total_tak_konsisten", "TOTAL != SCRIPLESS + SCRIP"),
        ("pct_di_luar_batas", "Persentase di luar (0, 100]"),
        ("emiten_kosong", "Nama emiten kosong"),
        ("emiten_tidak_seragam", "Nama emiten tidak seragam antar-baris satu kode"),
        ("investor_kosong", "Nama investor kosong"),
    ]:
        n = len(issues.get(key, []))
        status = "OK" if n == 0 else f"**{n} masalah**"
        add(f"- {label}: {status}")
        if n:
            ok = False
            for msg in issues[key][:5]:
                add(f"    - {msg}")
            if n > 5:
                add(f"    - ... (+{n-5} lagi)")
    add("")

    if baseline_rows is not None:
        base_codes = {r["Kode"] for r in baseline_rows}
        new_codes = sorted(set(codes) - base_codes)
        gone_codes = sorted(base_codes - set(codes))
        add("## Perbandingan dengan CSV baseline")
        add(f"- Emiten baseline: {len(base_codes)} | Excel: {len(codes)}")
        add(f"- Emiten baru: {new_codes or 'tidak ada'}")
        add(f"- Emiten hilang: {gone_codes or 'tidak ada'}")
        add("")

    add(f"## Kesimpulan: {'LULUS — siap dipakai Tahap 2' if ok else 'ADA MASALAH — lihat rincian di atas'}")
    return "\n".join(lines), ok


def main(argv=None):
    ap = argparse.ArgumentParser(description="Parser Excel KSEI kepemilikan 1%")
    ap.add_argument("excel", help="path file .xlsx laporan 1%")
    ap.add_argument("-o", "--output", required=True, help="path CSV keluaran")
    ap.add_argument("--sheet", help="nama sheet (default: sheet pertama)")
    ap.add_argument("--meta", help="path JSON metadata keluaran")
    ap.add_argument("--report", help="path laporan validasi (markdown)")
    ap.add_argument("--baseline-csv", help="CSV baseline situs untuk cross-check")
    args = ap.parse_args(argv)

    records, issues, stats = parse_excel(args.excel, args.sheet)
    if not records:
        print("FATAL: tidak ada baris data terbaca.", file=sys.stderr)
        return 2

    dates = sorted({r["date"] for r in records})

    with open(args.output, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLS)
        w.writeheader()
        for r in records:
            w.writerow({k: r[k] for k in COLS})

    if args.meta:
        meta = {
            "sumber_excel": args.excel,
            "tanggal_mentah": dates,
            "jumlah_baris": len(records),
            "jumlah_emiten": len({r["share_code"] for r in records}),
        }
        with open(args.meta, "w", encoding="utf-8") as f:
            json.dump(meta, f, ensure_ascii=False, indent=2)

    baseline_rows = None
    if args.baseline_csv:
        with open(args.baseline_csv, encoding="utf-8-sig") as f:
            baseline_rows = list(csv.DictReader(f))

    report, ok = build_report(records, issues, stats, args.excel, baseline_rows)
    if args.report:
        with open(args.report, "w", encoding="utf-8") as f:
            f.write(report + "\n")
    print(report)
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())